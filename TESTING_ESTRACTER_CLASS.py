from ESTRACTER import Estracter_data
from openpyxl import load_workbook
estracter = Estracter_data('05_Mayo_2023')
#%%
#print(estracter.path)

wb = load_workbook('EXCEL_FILES/05_Mayo_2023.xlsx',data_only=True)
shet = wb['CELEPSA']

#%%
#ACCEDIENDO A UNA CELDA EN ESPECÍFICO
cell = shet['K11']
#ACCEDIENDO A LOS ATRIBUTOS DEL OBJETO CELDA
# VALOR DE LA CELDA
print(cell.value)
# POSICIONAMIENTO EN FILA
print(cell.row)
# POSICIONAMIENTO EN COLUMNA(LETRA) 
print(cell.column_letter)

#%%
# NUMERO DE FILAS DE UNA HOJA
print(shet.max_row)

#%%
# CONVERSION DE CARACTERES A CODIGO ASSCI Y VISCEVERSA
#LETRA 
print('K')
#IMPRIMIR EL CODIGO ASCCI DEL CARACTER INGRESADO
print(ord('K'))
#IMPRIMIR CARACTER DEL CODIGO INGRESADO
print(chr(75))
#%%
# PROBANDO LA FUNCIONALIDAD DE EXTRAER DATOS DE UNA HOJA DE EXCEL Y ALMACENARLO EN UN DF
dic_df = estracter.test_hoja()

#%%
FINAL_DF = estracter.EXTRACT_ALL_DATA_FROM_EXCEL()
#%%
import pandas as pd

dic_test = {}

dic_test['a'] = ['ñgds','gdsg','dgsbc','qetgh','zbcdsa']
dic_test['b'] = ['7861','135741','78799922','1230698','4879543']
dic_test['c'] = ['rrwr','rwbhxf','qwrwqr','cbnxc','welokj']
dic_test['d'] = ['nvcmm','twettgb','vxbx','cxbeg','rwgwr']
dic_test['e'] = [None for i in range(5)]


df = pd.DataFrame(dic_test)
#%%
is_important = estracter.test_hoja()

#%%
dic_test_2 = {}
df_2 = pd.DataFrame(dic_test_2)

df_fusion = pd.concat([df,df_2])

#%%
# HACIENDO USO DE LA FUNCION FILTER
arr = [None,7,7,None,None]
filter_list = filter(lambda item: item is not None, arr)
new_arr = list(filter_list)
print(new_arr)

#%%

import statistics
arr = [7,9,5,5,1,2,3]
print(statistics.mode(arr))


#%%

print(None +1)






















