from ESTRACTER import Estracter_data
from openpyxl import load_workbook
## APRENDIZAJES
#print(estracter.path)
# OBTENIENDO EN UNA VARIABLE UN OBJETO QUE CONTENDRÁ EL EXCEL
wb = load_workbook('EXCEL_FILES/05_Mayo_2023.xlsx',data_only=True)
# OBTENIENDO EN UNA VARIABLE UN OBJETO QUE OCNTENDRÁ LA HOJA
shet = wb['CELEPSA']

#ACCEDIENDO A UNA CELDA EN ESPECÍFICO
cell = shet['K11']
#ACCEDIENDO A LOS ATRIBUTOS DEL OBJETO CELDA
# VALOR DE LA CELDA
print(cell.value)
# POSICIONAMIENTO EN FILA
print(cell.row)
# POSICIONAMIENTO EN COLUMNA(LETRA) 
print(cell.column_letter)
#%%
# NUMERO DE FILAS DE UNA HOJA
print(shet.max_row)
#%%
# CONVERSION DE CARACTERES A CODIGO ASSCI Y VISCEVERSA
#LETRA 
print('K')
#IMPRIMIR EL CODIGO ASCCI DEL CARACTER INGRESADO
print(ord('K'))
#IMPRIMIR CARACTER DEL CODIGO INGRESADO
print(chr(75))
#%%
# HACIENDO USO DE LA FUNCION FILTER
arr = [None,7,7,None,None]
filter_list = filter(lambda item: item is not None, arr)
new_arr = list(filter_list)
print(new_arr)
