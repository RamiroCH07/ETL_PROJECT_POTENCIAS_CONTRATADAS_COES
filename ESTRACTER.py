# CLASE QUE TENDRA FUNCIONALIDADES RELACIONADAS A EXTRAER DATOS DE ARCHVISO EXCEL
from openpyxl import load_workbook
import pandas as pd
from difflib import SequenceMatcher
import statistics
#%%
class Estracter_data:
    def __init__(self,cod_file):
        self.path = f'EXCEL_FILES/{cod_file}.xlsx'
        self.wb = load_workbook(self.path,data_only=True)
        self.camp_names = [
            'CODIGO DE RETIRO',
            'SUMINISTRADOR',
            'CLIENTE (1)',
            'TIPO DE CONTRATO',
            'TIPO DE USUARIO',
            'BARRA DE TRANSFERENCIA (2)',
            'FECHA INICIO',
            'FECHA FIN',
            'PUNTO(S) DE SUMINISTRO (3)',
            'DATOS DE LA POTENCIA CONTRATADA FIJA (MW)',
            'DATOS DE LA POTENCIA CONTRATADA VARIABLE (MW)',
            'COMENTARIO / OBSERVACIÓN'
            ]
        self.dic_lambdas = {
            self.camp_names[0]: lambda cadena: SequenceMatcher(a = self.camp_names[0], b = cadena.upper()).ratio(),
            self.camp_names[1]: lambda cadena: SequenceMatcher(a = self.camp_names[1], b = cadena.upper()).ratio(),
            self.camp_names[2]: lambda cadena: SequenceMatcher(a = self.camp_names[2], b = cadena.upper()).ratio(),
            self.camp_names[3]: lambda cadena: SequenceMatcher(a = self.camp_names[3], b = cadena.upper()).ratio(),
            self.camp_names[4]: lambda cadena: SequenceMatcher(a = self.camp_names[4], b = cadena.upper()).ratio(),
            self.camp_names[5]: lambda cadena: SequenceMatcher(a = self.camp_names[5], b = cadena.upper()).ratio(),
            self.camp_names[6]: lambda cadena: SequenceMatcher(a = self.camp_names[6], b = cadena.upper()).ratio(),
            self.camp_names[7]: lambda cadena: SequenceMatcher(a = self.camp_names[7], b = cadena.upper()).ratio(),
            self.camp_names[8]: lambda cadena: SequenceMatcher(a = self.camp_names[8], b = cadena.upper()).ratio(),
            self.camp_names[9]: lambda cadena: SequenceMatcher(a = self.camp_names[9], b = cadena.upper()).ratio(),
            self.camp_names[10]: lambda cadena: SequenceMatcher(a = self.camp_names[10], b = cadena.upper()).ratio(),
           self.camp_names[11]: lambda cadena: SequenceMatcher(a = self.camp_names[11], b = cadena.upper()).ratio()
            }
        
        self.wb = load_workbook(self.path,data_only=True)
        self.sheet_names = set(self.wb.sheetnames) - {'BASE','base','Base','ENEL DISTRIBUCIÓN'} 
      
    
    ## FUNCIONALIDAD QUE NOS RETORNA VERDADERO O FALSO SEGÚN SI LO QUE ENCUENTRE EN LA HOJA ES INFORMACIÓN QUE SE 
    ## DEBERÍA EXTRAER
    def _is_relevant_data(self,hoja):
        for row in hoja.iter_rows(min_row = 1, max_row = 30):
            for cell in row:
                #print(cell.value)
                if str(type(cell.value)) == "<class 'str'>":
                    port_sim_cod = self.dic_lambdas[self.camp_names[0]](cell.value)
                    port_sim_cli = self.dic_lambdas[self.camp_names[2]](cell.value)
                    port_sim_cont = self.dic_lambdas[self.camp_names[3]](cell.value)
                    port_sim_use = self.dic_lambdas[self.camp_names[4]](cell.value)
                    if port_sim_cod > 0.77 or port_sim_cli > 0.77 or port_sim_cont > 0.77 or port_sim_use > 0.77:
                        return True
        return False 

        
    def _identify_coordinates_cell(self,hoja,camp_name,key_name = None):
        for row in hoja.iter_rows(min_row = 1, max_row = 30):
            for cell in row:
                if str(type(cell.value)) == "<class 'str'>":
                    cell_value = str(cell.value)
                    port_sim = self.dic_lambdas[camp_name](cell_value)
                    if port_sim > 0.77:
                        if key_name is not None:
                            for word in cell_value.split():
                                #print(word.upper(),key_name)
                                sim_per = SequenceMatcher(a = key_name , b = word.upper()).ratio()   
                                if sim_per > 0.9:
                                    return(int(cell.row),cell.column_letter),True
                                
                        else:            
                            return (int(cell.row),cell.column_letter),True
        return (None,None),False
    
    
    
    def _get_data_from_column(self,hoja,start,end,column_letter):
        data_list = []
        for i in range(start,end+1):
            ubi = f'{column_letter}{i}'
            cell = hoja[ubi]
            if type(cell).__name__ == 'MergedCell' and len(data_list) > 0:
                ult_pos = len(data_list) - 1
                data_list.append(data_list[ult_pos])
            else:
                data_list.append(str(cell.value).strip())
        return data_list
            
        

    def _get_start_row(self,hoja):
        coordinates_cod = self._identify_coordinates_cell(hoja,self.camp_names[0])
        coordinates_cli = self._identify_coordinates_cell(hoja,self.camp_names[2])
        coordinates_cont = self._identify_coordinates_cell(hoja,self.camp_names[3])
        coordinates_use = self._identify_coordinates_cell(hoja,self.camp_names[4])
        coordinates_fini = self._identify_coordinates_cell(hoja,self.camp_names[6])
        coordinates_ffin = self._identify_coordinates_cell(hoja,self.camp_names[7])
        
        nums_fil = [coordinates_cod[0][0], 
                    coordinates_cli[0][0], 
                    coordinates_cont[0][0],
                    coordinates_use[0][0],
                    coordinates_fini[0][0],
                    coordinates_ffin[0][0]]
        nums_fil = list(set(nums_fil))
        filtered_ = filter(lambda item: item is not None, nums_fil)
        nums_fil = list(filtered_)
        start = statistics.mode(nums_fil) + 1
        return start
    
    def _get_end_row(self,hoja):
        return int(hoja.max_row)
    
    def _estract_data_x_sheet_to_df(self,hoja,name_sheet):
        dic_df = {}
        if self._is_relevant_data(hoja):
            dic_fulled = {}
            start = self._get_start_row(hoja)
            end = self._get_end_row(hoja)
            for name in self.camp_names:
                if name == self.camp_names[1]:
                    exist_camp = True
                if name == self.camp_names[9]:
                    coor,exist_camp = self._identify_coordinates_cell(hoja, name,key_name = 'FIJA')
                elif name == self.camp_names[10]:
                    coor,exist_camp = self._identify_coordinates_cell(hoja, name,key_name = 'VARIABLE')
                else:
                    coor,exist_camp = self._identify_coordinates_cell(hoja, name)     
                if exist_camp:
                    #cell = hoja[f'{coor[1]}{coor[0]}']
                    dic_fulled[name] = True
                    if name == self.camp_names[9] or name == self.camp_names[10]:
                        column_letter_ascci_1 = int(ord(coor[1]))
                        column_letter_ascci_2 = column_letter_ascci_1 + 1
                        column_letter_ascci_3 = column_letter_ascci_2 + 1
                        letter_column_total = coor[1]
                        letter_column_hp = chr(column_letter_ascci_2)
                        letter_column_hfp = chr(column_letter_ascci_3)
                        if name == self.camp_names[9]:
                            dic_df['POTENCIA_CONTRATADA_FIJA_TOTAL(MW)'] = self._get_data_from_column(hoja,start,end,letter_column_total)
                            dic_df['POTENCIA_CONTRATADA_FIJA_HORA_PUNTA'] = self._get_data_from_column(hoja,start,end,letter_column_hp)
                            dic_df['POTENCIA_CONTRATADA_FIJA_HORA_FUERA_PUNTA'] = self._get_data_from_column(hoja,start,end,letter_column_hfp)
                        else:
                            dic_df['POTENCIA_CONTRATADA_VARIABLE_TOTAL(MW)'] = self._get_data_from_column(hoja,start,end,letter_column_total)
                            dic_df['POTENCIA_CONTRATADA_VARIABLE_HORA_PUNTA'] = self._get_data_from_column(hoja,start,end,letter_column_hp)
                            dic_df['POTENCIA_CONTRATADA_VARIABLE_HORA_FUERA_PUNTA'] = self._get_data_from_column(hoja,start,end,letter_column_hfp)
                    elif name == self.camp_names[1]:
                        num_rows = end-start+1 
                        dic_df[name] = [name_sheet for _ in range(num_rows)]
                        
                    else:
                        dic_df[name] = self._get_data_from_column(hoja, start, end, coor[1])
                
                else:
                    num_rows = end-start+1
                    dic_df[name] = [None for i in range(num_rows)]
            #df = pd.DataFrame(dic_df)
            #return df
        df = pd.DataFrame(dic_df)
        return df
    
    def EXTRACT_ALL_DATA_FROM_EXCEL(self):
        columns_names = [
            'CODIGO DE RETIRO',
            'SUMINISTRADOR',
            'CLIENTE (1)',
            'TIPO DE CONTRATO',
            'TIPO DE USUARIO',
            'BARRA DE TRANSFERENCIA (2)',
            'FECHA INICIO',
            'FECHA FIN',
            'PUNTO(S) DE SUMINISTRO (3)',
            'POTENCIA_CONTRATADA_FIJA_TOTAL(MW)',
            'POTENCIA_CONTRATADA_FIJA_HORA_PUNTA',
            'POTENCIA_CONTRATADA_FIJA_HORA_FUERA_PUNTA',
            'POTENCIA_CONTRATADA_VARIABLE_TOTAL(MW)',
            'POTENCIA_CONTRATADA_VARIABLE_HORA_PUNTA',
            'POTENCIA_CONTRATADA_VARIABLE_HORA_FUERA_PUNTA',
            'COMENTARIO / OBSERVACIÓN']
        df = pd.DataFrame(columns = columns_names)
        for name in self.sheet_names:
            print('EXTRAYENDO DATOS DE HOJA:'+' '+name)
            df = pd.concat([df,self._estract_data_x_sheet_to_df(self.wb[name],name)])
        return df
            
            
            
    def test_hoja(self):
        sheet = self.wb['AGROAURORA']
        return self._is_relevant_data(sheet)
                
            
                
                    
                
                
                
                
        
        
    
    
    
    
                    
    
    
    
    
        
    

        
   
        
